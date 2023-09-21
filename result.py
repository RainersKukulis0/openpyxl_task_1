from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
max_row=ws.max_row
for i in range(2,max_row+1):
    rate=(ws["B"+str(i)].value)
    hours=(ws["C"+str(i)].value)
    if(type(hours)!=str and type(rate)!=str):
        salary=float(hours)*float(rate)
        if(salary>3000):
          total+=1
print(total)
wb.close()